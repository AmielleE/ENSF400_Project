import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter import Tk, filedialog


def export_to_excel(deadline_df: pd.DataFrame, output_path: str | None = None) -> None:
    """
    Takes a DataFrame with columns:
        - classname
        - assignment_name
        - due_date

    Creates an Excel file that:
        - sorts deadlines by due date ascending
        - adds a Done column
        - color-codes rows by class
        - allows filtering/sorting in Excel

    If no output_path is provided, the user is asked where to save the file.
    """

    # If no output path is given, open a save dialog so the user can choose
    if output_path is None:
        root = Tk()
        root.withdraw()  # Hide the small main tkinter window

        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="deadlines.xlsx",
            title="Save Excel File As"
        )

        root.destroy()

        # If the user cancels the save dialog, stop the function
        if not output_path:
            print("Save cancelled.")
            return

    # Make a copy so the original DataFrame is not modified
    df = deadline_df.copy()

    # Make sure required columns exist
    required_columns = ["classname", "assignment_name", "due_date"]
    for column in required_columns:
        if column not in df.columns:
            raise ValueError(f"Missing required column: {column}")

    # Convert due_date strings to actual datetime values so sorting works properly
    df["due_date"] = pd.to_datetime(df["due_date"], format="%m/%d/%Y", errors="coerce")

    # Sort by due date ascending
    df = df.sort_values(by="due_date", ascending=True)

    # Add a Done column with default value "Not Done"
    df["done_status"] = "Not Done"

    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Deadlines"

    # Define final column order
    columns = ["classname", "assignment_name", "due_date", "done_status"]

    # Write headers
    for col_num, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Assign a fill color to each class
    class_colors = {}
    color_palette = [
        "FFF2CC",  # light yellow
        "D9EAD3",  # light green
        "DDEBF7",  # light blue
        "F4CCCC",  # light red
        "EAD1DC",  # light purple
        "FCE5CD",  # light orange
        "D0E0E3",  # light teal
        "E2F0D9"   # pale green
    ]

    unique_classes = df["classname"].dropna().unique()
    for i, class_name in enumerate(unique_classes):
        class_colors[class_name] = color_palette[i % len(color_palette)]

    # Write data rows
    for row_num, (_, row) in enumerate(df.iterrows(), start=2):
        class_name = row["classname"]
        row_fill_color = class_colors.get(class_name, "FFFFFF")

        values = [
            row["classname"],
            row["assignment_name"],
            row["due_date"],
            row["done_status"]
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)

            # Apply row color based on class
            cell.fill = PatternFill(fill_type="solid", fgColor=row_fill_color)

            # Format date column properly
            if columns[col_num - 1] == "due_date" and pd.notna(value):
                cell.number_format = "MM/DD/YYYY"

            # Align cells
            cell.alignment = Alignment(vertical="center")

    # Create an Excel table so the user can filter and sort by class or anything else
    table_end_row = ws.max_row
    table_end_col = ws.max_column
    table_ref = f"A1:{get_column_letter(table_end_col)}{table_end_row}"

    table = Table(displayName="DeadlineTable", ref=table_ref)
    table_style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False
    )
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    # Save the workbook
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")